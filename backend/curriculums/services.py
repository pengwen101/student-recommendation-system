from backend.curriculums import cypher as curriculum_cypher
from backend.curriculums import template as curriculum_template
from fastapi import HTTPException, UploadFile
from typing import List
import openpyxl
import uuid


async def read_curriculum(version_id: str):
    return await curriculum_cypher.read_curriculum(version_id)


async def study_level_exists(study_level_id: str):
    return await curriculum_cypher.study_level_exists(study_level_id)


async def read_curriculum_versions():
    return await curriculum_cypher.read_curriculum_versions()


async def read_questions(batch_id: str | None = None):
    if batch_id:
        questions = await curriculum_cypher.read_questions_by_batch(batch_id)
    else:
        questions = await curriculum_cypher.read_questions()
    return questions


async def get_batch_info(version_id: str):
    return await curriculum_cypher.get_batch_info_for_version(version_id)


async def generate_template():
    return curriculum_template.generate_template()


async def create_curriculum(file: UploadFile, batch_id: str):
    wb = openpyxl.load_workbook(file.file)

    # --- Parse Sheet 1: CPL & SubCPL ---
    ws1 = wb["CPL & SubCPL"]
    sheet1_rows = []
    for row in ws1.iter_rows(min_row=2, values_only=True):
        cpl_code, cpl_name, sub_cpl_code, sub_cpl_name, weight, quality_name = row
        if not cpl_code:
            continue
        sheet1_rows.append({
            "cpl_code": str(cpl_code).strip(),
            "cpl_name": str(cpl_name or "").strip(),
            "sub_cpl_code": str(sub_cpl_code).strip(),
            "sub_cpl_name": str(sub_cpl_name or "").strip(),
            "weight": float(weight or 1),
            "quality_name": str(quality_name or "").strip(),
        })

    if not sheet1_rows:
        raise HTTPException(status_code=400, detail="Sheet 1 (CPL & SubCPL) is empty")

    # Deduplicate CPLs — code becomes "CPL{num}"
    cpl_map: dict[str, dict] = {}
    for r in sheet1_rows:
        code = r["cpl_code"]
        if code not in cpl_map:
            display_code = f"CPL{code}"
            cpl_map[code] = {"cpl_id": str(uuid.uuid4()), "code": display_code, "name": r["cpl_name"]}

    # Deduplicate SubCPLs, track parent CPL — code becomes "CPL{x}S{y}"
    subcpl_map: dict[str, dict] = {}
    for r in sheet1_rows:
        code = r["sub_cpl_code"]
        if code not in subcpl_map:
            parts = code.split(".")
            cpl_num = parts[0]
            sub_num = parts[1] if len(parts) > 1 else "1"
            display_code = f"CPL{cpl_num}S{sub_num}"
            subcpl_map[code] = {
                "sub_cpl_id": str(uuid.uuid4()),
                "code": display_code,
                "name": r["sub_cpl_name"],
                "cpl_id": cpl_map[r["cpl_code"]]["cpl_id"],
            }

    # Deduplicate Qualities by name (many-to-many with SubCPL) — globally sequential "Q001", "Q002"...
    quality_map: dict[str, dict] = {}
    quality_links: list[dict] = []
    quality_global_counter = 0
    for r in sheet1_rows:
        name = r["quality_name"]
        if name not in quality_map:
            quality_global_counter += 1
            q_code = f"Q{quality_global_counter:03d}"
            quality_map[name] = {
                "quality_id": str(uuid.uuid4()),
                "code": q_code,
                "name": name,
            }
        quality_links.append({
            "sub_cpl_id": subcpl_map[r["sub_cpl_code"]]["sub_cpl_id"],
            "quality_id": quality_map[name]["quality_id"],
            "weight": r["weight"],
        })

    # --- Parse Sheet 2: Indikator & Pertanyaan ---
    ws2 = wb["Indikator & Pertanyaan"]
    sheet2_rows = []
    for row in ws2.iter_rows(min_row=2, values_only=True):
        quality_name, indicator_name, question_name, flipped, scale_label = row
        if not quality_name:
            continue
        sheet2_rows.append({
            "quality_name": str(quality_name).strip(),
            "indicator_name": str(indicator_name or "").strip(),
            "question_name": str(question_name or "").strip(),
            "flipped": _parse_flipped(flipped),
            "scale_label": str(scale_label or "").strip(),
        })

    # Validate quality names in Sheet 2 exist in Sheet 1
    s1_quality_names = {r["quality_name"] for r in sheet1_rows}
    for r in sheet2_rows:
        if r["quality_name"] not in s1_quality_names:
            raise HTTPException(
                status_code=400,
                detail=f"Quality '{r['quality_name']}' in Sheet 2 not found in Sheet 1",
            )

    # Build quality name → quality_id mapping from Sheet 1
    quality_name_to_id = {name: q["quality_id"] for name, q in quality_map.items()}

    # Deduplicate Indicators — globally sequential "I001", "I002"...
    indicator_map: dict[str, dict] = {}
    indicator_global_counter = 0
    for r in sheet2_rows:
        q_id = quality_name_to_id[r["quality_name"]]
        key = f"{q_id}|{r['indicator_name']}"
        if key not in indicator_map:
            indicator_global_counter += 1
            ind_code = f"I{indicator_global_counter:03d}"
            indicator_map[key] = {
                "indicator_id": str(uuid.uuid4()),
                "code": ind_code,
                "name": r["indicator_name"],
                "quality_id": q_id,
            }

    # Deduplicate Questions — globally sequential "QS001", "QS002"...
    question_map: dict[str, dict] = {}
    question_global_counter = 0
    for r in sheet2_rows:
        q_id = quality_name_to_id[r["quality_name"]]
        ind_key = f"{q_id}|{r['indicator_name']}"
        ind = indicator_map.get(ind_key)
        if not ind:
            continue
        key = f"{ind['indicator_id']}|{r['question_name']}|{r['scale_label']}|{r['flipped']}"
        if key not in question_map:
            question_global_counter += 1
            qs_code = f"QS{question_global_counter:03d}"
            question_map[key] = {
                "question_id": str(uuid.uuid4()),
                "code": qs_code,
                "name": r["question_name"],
                "scale_label": r["scale_label"],
                "flipped": r["flipped"],
                "indicator_id": ind["indicator_id"],
            }

    # --- Create in Neo4j ---
    max_id = await curriculum_cypher.get_max_curriculum_version_id()
    version_id = str(int(max_id) + 1)

    await curriculum_cypher.create_curriculum_version(version_id)
    await curriculum_cypher.batch_create_cpls(version_id, list(cpl_map.values()))
    await curriculum_cypher.batch_create_subcpls(list(subcpl_map.values()))
    await curriculum_cypher.batch_create_qualities(list(quality_map.values()))
    await curriculum_cypher.batch_link_qualities(quality_links)
    await curriculum_cypher.batch_create_indicators(list(indicator_map.values()))
    await curriculum_cypher.batch_create_questions(list(question_map.values()))
    await curriculum_cypher.link_version_to_batch(version_id, batch_id)

    return await curriculum_cypher.read_curriculum(str(version_id))


async def delete_curriculum_version(version_id: str):
    version_exists = await curriculum_cypher.read_curriculum_versions()
    if not any(v.get("curriculum_version_id") == version_id for v in version_exists):
        raise HTTPException(status_code=404, detail=f"Curriculum version {version_id} not found")

    batch_info = await curriculum_cypher.get_batch_info_for_version(version_id)
    total_students = sum(b.get("student_count", 0) for b in batch_info)
    if total_students > 0:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot delete curriculum version {version_id}: used by {total_students} student(s)"
        )
    await curriculum_cypher.delete_curriculum_version(version_id)


def _parse_flipped(val):
    if isinstance(val, bool):
        return val
    if isinstance(val, str):
        return val.strip().upper() in ("TRUE", "1", "YES")
    return False