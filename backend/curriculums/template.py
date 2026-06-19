from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment


def generate_template() -> BytesIO:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "CPL & SubCPL"
    ws1.append([
        "Capaian Pembelajaran Lulusan",
        "Deskripsi Capaian Pembelajaran Lulusan",
        "Sub-Capaian Pembelajaran Lulusan",
        "Deskripsi Sub-Capaian Pembelajaran Lulusan",
        "Bobot",
        "Kualitas",
    ])
    _style_header(ws1)
    _auto_width(ws1, [30, 50, 30, 50, 10, 40])

    ws2 = wb.create_sheet("Indikator & Pertanyaan")
    ws2.append([
        "Kualitas",
        "Indikator",
        "Pertanyaan",
        "Flipped",
        "Label Jawaban",
    ])
    _style_header(ws2)
    _auto_width(ws2, [40, 40, 50, 10, 40])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _style_header(ws):
    font = Font(bold=True)
    for cell in ws[1]:
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
